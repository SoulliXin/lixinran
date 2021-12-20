# -*- codeing = utf-8 -*-
# @time : 2021/12/8 10:41
# @Author : LiXinRan
# @File : excel_util.py
# @Software: PyCharm
import openpyxl
import time
import os
import sys
base_path = os.path.abspath(os.path.dirname(__file__)).split('util')[0]
sys.path.append(base_path)

class ExcelUtil:
    def __init__(self,file_path=None,index=None):
        if  file_path == None:
            self.file_path = base_path + 'config/case.xlsx'
        else:
            self.file_path =file_path
        if index == None:
            index = 0
        self.op = openpyxl.load_workbook(self.file_path)
        self.sheetname = self.op.sheetnames[index]
        self.data = self.op[self.sheetname]

    def get_lines(self):
        rows = self.data.max_row
        if rows >=1:
            return rows
        return None

    def get_value(self,row,col):
        if self.get_lines()>=row:
            value = self.data.cell(row,col).value
            return value
        return None

    def get_row_data(self,row):
        row_list = []
        for i in self.data[row]:
            row_list.append(i.value)
        return row_list

    def get_data(self):
        result = []
        for i in range(self.get_lines()-1):
            data = self.get_row_data(i+2)
            result.append(data)
        return result


    def write_excel_data(self,row,col,value):
        wb = self.op
        wr = wb.active
        wr.cell(row,col,value)
        wb.save(base_path + 'config/case.xlsx')


if __name__ == '__main__':
    ex = ExcelUtil('c:/ECshopUItest/config/case.xlsx',1)
    data = ex.get_data()
    print(data)




